import os
import pandas as pd
from multiprocessing import Process, Queue
from time import sleep 
from collections.abc import Callable
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv
from .load_data import load_datasets_with_annotations as loading


load_dotenv()

def check_if_sheet_exists_and_delete(q: Queue) -> None:
    delete_flag: bool = False
    path_to_file: str = str(q.get())
    file_name: str = str(q.get())
    sheet_name: str = str(q.get())
    full_path: str = os.path.join(path_to_file, file_name)
    try:
        wb = load_workbook(full_path)
        if (sheet_name in wb.sheetnames and len(wb.sheetnames) == 1):
            delete_flag = True
        else:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            wb.save(full_path)
    except FileNotFoundError:
        print(f"File not found: {full_path}")
    except PermissionError:
        print(f"Permission denied while accessing {full_path}")
    except Exception as e:
        print(f"An unexpected error occurred while processing {full_path}: {e}")
    finally:
        wb.close()
        q.put(delete_flag)

def save_to_excel(
    local_data: pd.DataFrame,
    formatter: Callable[[Workbook, str], None] = None,
    sheet_name: str = "Sheet",
    name_xlsx: str = os.getenv("OUTPUT_EXCEL_NAME_WITHOUT_ANNOTATIONS"),
):
    mode: str = "w"
    if os.path.exists(name_xlsx):
        q: Queue = Queue()
        q.put(os.getcwd())
        q.put(name_xlsx)
        q.put(sheet_name)
        process = Process(
            target=check_if_sheet_exists_and_delete,
            args=(q,),
        )
        process.start()
        process.join()
        while process.is_alive():
            sleep(1)
        process.close()
        if not q.empty():
            if bool (q.get()):
                full_file_path: str = os.path.join(os.getcwd(), name_xlsx)
                os.remove(full_file_path)
            else:
                mode = "r+"
    with pd.ExcelWriter(name_xlsx, mode=mode) as writer:
        local_data.to_excel(writer, index=False, sheet_name=sheet_name)
    if formatter:
        wb = load_workbook(name_xlsx)
        formatter(wb, sheet_name)
        wb.save(name_xlsx)
        wb.close()

def load_datasets_add_line_counter() -> dict[str, list]:
    datasets: dict[str, list] = loading()
    for key in datasets.keys():
        current_dataset = datasets[key]
        for i, item in enumerate(current_dataset, start=1):
            item = {"linecounter": i, **item}
            current_dataset[i - 1] = item
    return datasets


def prepaire_excel_data() -> pd:
    current_directory = os.getcwd()
    suffix = "\\src"
    directory_excel = (
        current_directory[: -len(suffix)]
        if current_directory.endswith(suffix)
        else current_directory
    )
    directory_excel += "\\Datasets\\Evaluation_v4.xlsx"
    excel_data = pd.read_excel(
        directory_excel,
        usecols=lambda column: "Item" not in column,
        skiprows=range(1),
        sheet_name="Ground Truth",
    )
    excel_data = excel_data.drop(
        columns=[
            "Unnamed: 0",
            "Unnamed: 12",
            "total",
            "main",
            "benefit",
            "total.1",
            "main.1",
            "benefit.1",
        ],
        axis=1,
    )
    excel_data = excel_data.rename(
        columns={
            "Main Part \nPartial": "Main Part Partial",
            "Main Part \nFull": "Main Part Full",
            "Benefit\nPartial": "Benefit Partial",
            "Benefit\nFull": "Benefit Full",
        }
    )
    excel_data.index += 1
    excel_data = excel_data.fillna("Empty")

    ## Map line number to user story number
    excel_data.insert(loc=4, column="Corresponding USID 1", value=0)
    excel_data.insert(loc=5, column="Corresponding USID 2", value=0)
    datasets = load_datasets_add_line_counter()
    for idx in range(len(excel_data)):
        redundant_pairs = excel_data.iat[idx, 1]
        parts = redundant_pairs.split("_")
        first_number = int(parts[2])
        second_number = int(parts[-1])
        project_number = f"#{excel_data.iat[idx, 0]}#".upper()
        project_data = datasets[project_number]
        for item in project_data:
            if item["linecounter"] == first_number:
                excel_data.iat[idx, 4] = item["id"]
            if item["linecounter"] == second_number:
                excel_data.iat[idx, 5] = item["id"]
    return excel_data

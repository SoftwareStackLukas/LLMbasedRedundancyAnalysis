import os
import pandas as pd
from collections.abc import Callable
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv
from functools import wraps

def load_env(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        load_dotenv()  # Load environment variables
        return func(*args, **kwargs)
    return wrapper

@load_env
def save_to_excel(local_data: pd.DataFrame, formatter: Callable[[Workbook, str], None] = None, sheet_name: str = "Sheet", name_xlsx: str = os.getenv("OUTPUT_EXCEL_NAME")):
    if os.path.exists(name_xlsx):
        try:
            wb = load_workbook(name_xlsx)
            if sheet_name in wb.sheetnames:
                try:
                    del wb[sheet_name]
                    wb.save(name_xlsx)
                finally:
                    if 'wb' in locals() and wb:
                        wb.close()
                        wb = None
                    wb = load_workbook(name_xlsx)    
            if (len(wb.sheetnames) >= 1):
                with pd.ExcelWriter(name_xlsx, mode='a') as writer:
                    local_data.to_excel(writer, index=False , sheet_name=sheet_name)
                if ('writer' in locals()):
                    writer = None
            else:
                with pd.ExcelWriter(name_xlsx, mode='w') as writer:
                    local_data.to_excel(writer, index=False , sheet_name=sheet_name)
                if ('writer' in locals()):
                    writer = None
        finally:
            if 'wb' in locals() and wb:
                wb.close()
                wb = None
    else:
        with pd.ExcelWriter(name_xlsx, mode='w') as writer:
                    local_data.to_excel(writer, index=False , sheet_name=sheet_name)
        if ('writer' in locals()):
            writer = None
    try:
        wb:Workbook = None
        if formatter:
            wb:Workbook  = load_workbook(name_xlsx)
            formatter(wb, sheet_name)
            wb.save(name_xlsx)
    finally:
        if 'wb' in locals() and wb:
            wb.close()
            wb = None
            
            
def load_datasets_add_line_counter() -> dict[str, list]:
    datasets: dict[str, list] = loading()
    for key in datasets.keys():
        current_dataset = datasets[key]
        for i, item in enumerate(current_dataset, start=1):
            item = {'linecounter': i, **item}
            current_dataset[i-1] = item
    return datasets